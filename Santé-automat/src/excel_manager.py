"""
Module de gestion des fichiers Excel (mise à jour, colorisation, graphiques).
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference
from config import OUTPUT_FILE


class ExcelManager:
    """Gestionnaire des opérations sur le fichier Excel."""
    
    def __init__(self, output_file=OUTPUT_FILE):
        self.output_file = output_file
    
    def update_excel(self, pdf_path, date, results, thresholds_manager, force=False):
        """
        Met à jour le fichier Excel avec les nouvelles données.
        
        Args:
            pdf_path: Chemin du fichier PDF source
            date: Date du prélèvement
            results: Dictionnaire des résultats extraits
            thresholds_manager: Instance du gestionnaire de seuils
            force: Force la ré-extraction même si déjà présente
        """
        if os.path.exists(self.output_file):
            df = pd.read_excel(self.output_file)
            if date in df.columns and not force:
                print(f"[INFO] Date {date} déjà extraite, passez --force pour ré-extraire")
                return
        else:
            df = pd.DataFrame()
        
        if "Paramètre" not in df.columns:
            df["Paramètre"] = list(results.keys())
        
        # Colonnes de base
        for base_col in ["Unité", "Min", "Max"]:
            if base_col not in df.columns:
                df[base_col] = None
        
        if date not in df.columns:
            df[date] = None
        
        # Mise à jour des seuils depuis les résultats
        thresholds_manager.update_from_results(results)
        
        for param, info in results.items():
            if param not in df["Paramètre"].values:
                df.loc[len(df)] = [param] + [None] * (len(df.columns) - 1)
            df.loc[df["Paramètre"] == param, date] = info["valeur"]
            
            # Renseigner Unité/Min/Max dans l'Excel si disponibles et absents
            if info.get("unité") is not None:
                df.loc[df["Paramètre"] == param, "Unité"] = info["unité"]
            if info.get("min") is not None:
                df.loc[df["Paramètre"] == param, "Min"] = info["min"]
            if info.get("max") is not None:
                df.loc[df["Paramètre"] == param, "Max"] = info["max"]
        
        # Sauvegarder les seuils mis à jour
        thresholds_manager.save_thresholds()
        
        # Réordonner les colonnes de dates par ordre chronologique
        df = self._reorder_columns(df)
        
        df.to_excel(self.output_file, index=False)
        print(f"[INFO] {os.path.basename(pdf_path)} -> {len(results)} résultats extraits pour la date {date}")
    
    def _reorder_columns(self, df):
        """Réordonne les colonnes avec les dates en ordre chronologique."""
        def _is_date_label(c):
            if c == "Paramètre":
                return False
            if c in ("Unité", "Min", "Max"):
                return False
            return not pd.isna(pd.to_datetime(c, dayfirst=True, errors="coerce"))
        
        date_cols = [c for c in df.columns if _is_date_label(c)]
        non_date_cols = [c for c in df.columns if c not in date_cols and c != "Paramètre"]
        # tri par date croissante
        sorted_dates = sorted(date_cols, key=lambda x: pd.to_datetime(x, dayfirst=True, errors="coerce"))
        ordered_cols = ["Paramètre", "Unité", "Min", "Max"] + sorted_dates + non_date_cols
        return df.reindex(columns=ordered_cols)
    
    def colorize_outliers(self, thresholds_manager):
        """Colorise les valeurs hors normes dans le fichier Excel."""
        if not os.path.exists(self.output_file):
            return
        
        wb = load_workbook(self.output_file)
        ws = wb.active
        
        green = PatternFill(start_color="AAFFAA", end_color="AAFFAA", fill_type="solid")
        red = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")
        
        # Repérer colonnes
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        idx_param = headers.index("Paramètre") + 1 if "Paramètre" in headers else 1
        idx_min = headers.index("Min") + 1 if "Min" in headers else None
        idx_max = headers.index("Max") + 1 if "Max" in headers else None
        
        # Colonnes de dates
        date_cols_idx = []
        for i, h in enumerate(headers, start=1):
            if h in ("Paramètre", "Unité", "Min", "Max"):
                continue
            try:
                if not pd.isna(pd.to_datetime(h, dayfirst=True, errors="coerce")):
                    date_cols_idx.append(i)
            except Exception:
                pass
        
        for r in range(2, ws.max_row + 1):
            param = ws.cell(row=r, column=idx_param).value
            # seuils préférentiels: colonnes Min/Max de la feuille, sinon JSON
            tmin = tmax = None
            if idx_min and idx_max:
                vmin = ws.cell(row=r, column=idx_min).value
                vmax = ws.cell(row=r, column=idx_max).value
                try:
                    tmin = None if vmin in (None, "") else float(str(vmin).replace(",", "."))
                    tmax = None if vmax in (None, "") else float(str(vmax).replace(",", "."))
                except Exception:
                    tmin = tmax = None
            if (tmin is None or tmax is None):
                threshold = thresholds_manager.get_threshold(param)
                tmin = threshold.get("min")
                tmax = threshold.get("max")
            
            if tmin is None or tmax is None:
                continue
            
            for c in date_cols_idx:
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue
                try:
                    val = float(str(cell.value).replace(",", "."))
                except Exception:
                    continue
                cell.fill = green if (tmin <= val <= tmax) else red
        
        wb.save(self.output_file)
        print("[INFO] Colorisation Excel terminée")
    
    def insert_charts(self):
        """Insère des graphiques d'évolution dans le fichier Excel."""
        if not os.path.exists(self.output_file):
            return
        
        # Charger workbook et feuille de données
        wb = load_workbook(self.output_file)
        ws_data = wb.active
        
        # Trouver/Créer la feuille "Graphiques"
        sheet_name = "Graphiques"
        if sheet_name in wb.sheetnames:
            # Supprimer et recréer pour éviter doublons
            del wb[sheet_name]
        ws_charts = wb.create_sheet(title=sheet_name)
        
        # Repérer les colonnes de dates (en-têtes ligne 1)
        headers = [ws_data.cell(row=1, column=col).value for col in range(1, ws_data.max_column + 1)]
        try:
            date_cols_idx = [
                i + 1 for i, h in enumerate(headers)
                if h not in ("Paramètre", "Unité", "Min", "Max")
                and not pd.isna(pd.to_datetime(h, dayfirst=True, errors="coerce"))
            ]
        except Exception:
            date_cols_idx = [i + 1 for i, h in enumerate(headers) if h not in ("Paramètre", "Unité", "Min", "Max")]
        
        # Pour chaque paramètre (ligne >=2), créer un graphique si au moins 2 valeurs
        chart_row = 1
        chart_col = 1
        charts_per_row = 2
        for r in range(2, ws_data.max_row + 1):
            param_name = ws_data.cell(row=r, column=1).value
            if not param_name:
                continue
            
            # Compter valeurs numériques présentes
            numeric_count = 0
            for c in date_cols_idx:
                v = ws_data.cell(row=r, column=c).value
                try:
                    if v is not None and str(v) != "" and float(str(v).replace(",", ".")) == float(str(v).replace(",", ".")):
                        numeric_count += 1
                except Exception:
                    pass
            if numeric_count < 2:
                continue
            
            # Construire chart
            chart = LineChart()
            chart.title = f"Évolution de {param_name}"
            chart.y_axis.title = "Valeur"
            chart.x_axis.title = "Date"
            
            # Données
            data_ref = Reference(ws_data, min_col=min(date_cols_idx), max_col=max(date_cols_idx), min_row=r, max_row=r)
            cat_ref = Reference(ws_data, min_col=min(date_cols_idx), max_col=max(date_cols_idx), min_row=1, max_row=1)
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(cat_ref)
            
            # Position dans la feuille charts
            cell_addr = ws_charts.cell(row=chart_row, column=chart_col).coordinate
            ws_charts.add_chart(chart, cell_addr)
            
            # Avancer position (grille 2 colonnes)
            if chart_col == 1:
                chart_col = 9  # environ 8 colonnes d'écart
            else:
                chart_col = 1
                chart_row += 15  # descendre d'un bloc
        
        wb.save(self.output_file)
        print("[INFO] Graphiques intégrés dans le fichier Excel (onglet 'Graphiques')")
