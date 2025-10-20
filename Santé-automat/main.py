import os
import re
import json
import pdfplumber
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
import argparse

load_dotenv()

# Valeurs par défaut si .env n'est pas défini
PDF_FOLDER = os.getenv("PDF_FOLDER", "pdfs")
OUTPUT_FILE = os.getenv("OUTPUT_FILE", "resultats.xlsx")
THRESHOLDS_FILE = os.getenv("THRESHOLDS_FILE", "seuils.json")

# ---------------------------
# Liste blanche des paramètres biologiques
# ---------------------------
PARAMS_WHITELIST = [
    "Hématies", "Hémoglobine", "Hématocrite", "V.G.M.", "T.C.M.H.",
    "C.C.M.H.", "Leucocytes", "Polynucléaires neutrophiles",
    "Polynucléaires éosinophiles", "Polynucléaires basophiles",
    "Lymphocytes", "Monocytes", "Plaquettes",
    "Créatinine", "Phosphatase alcaline", "ASAT (Transaminases TGO)",
    "ALAT (Transaminases TGP)", "GGT (Gamma Glutamyl Transpeptidase)",
    "FIB-4"
]

# ---------------------------
# Extraction des données du PDF
# ---------------------------
def extract_data_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    # Date de prélèvement
    date_match = re.search(r"Prélevé le (\d{2}[-−]\d{2}[-−]\d{4})", text)
    if date_match:
        date = date_match.group(1).replace("−","-")
        text = text.split(date_match.group(0), 1)[1]  # texte après la date
    else:
        date = "inconnue"

    results = {}
    pattern = re.compile(
        r"^\s*([A-Za-zÀ-ÿ\s\-\(\)/]+?)\s+([\d,\.]+)\s*([a-zA-Zµ%/]+)?\s*(?:\(([^)]+)\))?"
    )

    for line in text.split("\n"):
        line = line.strip()
        if not line or len(line) > 120:
            continue

        # on ne garde que les paramètres whitelistés
        for param in PARAMS_WHITELIST:
            if line.startswith(param):
                match = pattern.match(line)
                if match:
                    val_str = match.group(2).replace(",", ".")
                    try:
                        value = float(val_str)
                        interval = match.group(4) or ""
                        # si intervalle = "min-max", on peut stocker les valeurs pour yerr
                        low, high = None, None
                        if interval and "-" in interval:
                            parts = interval.replace("−","-").split("-")
                            try:
                                low = float(parts[0].replace(",", "."))
                                high = float(parts[1].replace(",", "."))
                            except:
                                pass
                        results[param] = {
                            "valeur": value,
                            "unité": match.group(3) or "",
                            "intervalle": interval,
                            "min": low,
                            "max": high
                        }
                    except ValueError:
                        continue
                break  # ne pas matcher plusieurs fois

    return date, results

# ---------------------------
# Seuils: chargement / sauvegarde
# ---------------------------
def load_thresholds():
    thresholds = {}
    if not THRESHOLDS_FILE or not os.path.exists(THRESHOLDS_FILE):
        return thresholds

    _, ext = os.path.splitext(THRESHOLDS_FILE)
    try:
        if ext.lower() in [".json"]:
            with open(THRESHOLDS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            # normaliser {param: {min: x, max: y}}
            for k, v in data.items():
                try:
                    thresholds[k] = {
                        "min": None if v.get("min") is None else float(str(v.get("min")).replace(",", ".")),
                        "max": None if v.get("max") is None else float(str(v.get("max")).replace(",", ".")),
                    }
                except Exception:
                    continue
        elif ext.lower() in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
            df_thresh = pd.read_excel(THRESHOLDS_FILE, engine="openpyxl")
            for _, row in df_thresh.iterrows():
                try:
                    pname = str(row["Paramètre"]) if "Paramètre" in row else None
                    if not pname:
                        continue
                    vmin = row["min"] if "min" in row else None
                    vmax = row["max"] if "max" in row else None
                    thresholds[pname] = {
                        "min": None if vmin is None or pd.isna(vmin) else float(vmin),
                        "max": None if vmax is None or pd.isna(vmax) else float(vmax),
                    }
                except Exception:
                    continue
        else:
            print(f"[WARN] Extension de THRESHOLDS_FILE non supportée: {ext}")
    except Exception as e:
        print(f"[WARN] Impossible de lire THRESHOLDS_FILE : {e}")

    return thresholds


def save_thresholds(thresholds: dict):
    if not THRESHOLDS_FILE:
        return
    os.makedirs(os.path.dirname(THRESHOLDS_FILE) or ".", exist_ok=True)
    _, ext = os.path.splitext(THRESHOLDS_FILE)
    try:
        if ext.lower() == ".json":
            with open(THRESHOLDS_FILE, "w", encoding="utf-8") as f:
                json.dump(thresholds, f, ensure_ascii=False, indent=2)
        else:
            # Sauvegarde JSON par défaut même si extension non-json
            with open(THRESHOLDS_FILE + ".json", "w", encoding="utf-8") as f:
                json.dump(thresholds, f, ensure_ascii=False, indent=2)
            print(f"[INFO] Seuils enregistrés dans {THRESHOLDS_FILE + '.json'}")
    except Exception as e:
        print(f"[WARN] Échec d'enregistrement des seuils: {e}")

# ---------------------------
# Mise à jour Excel
# ---------------------------
def update_excel(pdf_path, date, results, force=False):
    if os.path.exists(OUTPUT_FILE):
        df = pd.read_excel(OUTPUT_FILE)
        if date in df.columns and not force:
            print(f"[INFO] Date {date} déjà extraite, passez --force pour ré-extraire")
            return
    else:
        df = pd.DataFrame()

    if "Paramètre" not in df.columns:
        df["Paramètre"] = list(results.keys())

    if date not in df.columns:
        df[date] = None

    # Charger seuils existants et les compléter à partir des PDF (si min/max présents)
    thresholds = load_thresholds()

    for param, info in results.items():
        if param not in df["Paramètre"].values:
            df.loc[len(df)] = [param] + [None]*(len(df.columns)-1)
        df.loc[df["Paramètre"]==param, date] = info["valeur"]

        # Enregistrer automatiquement un seuil si manquant et détecté
        if info.get("min") is not None and info.get("max") is not None:
            if param not in thresholds:
                thresholds[param] = {"min": info["min"], "max": info["max"]}

    # Sauvegarder les seuils mis à jour (si quelque chose a été ajouté)
    save_thresholds(thresholds)

    df.to_excel(OUTPUT_FILE, index=False)
    print(f"[INFO] {os.path.basename(pdf_path)} -> {len(results)} résultats extraits pour la date {date}")

# ---------------------------
# Colorisation Excel
# ---------------------------
def colorize_outliers():
    if not os.path.exists(OUTPUT_FILE):
        return

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    green = PatternFill(start_color="AAFFAA", end_color="AAFFAA", fill_type="solid")
    red = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")

    # Charger les seuils (JSON ou Excel)
    thresholds = load_thresholds()

    for row in ws.iter_rows(min_row=2, min_col=2):
        param = ws.cell(row=row[0].row, column=1).value
        if param in thresholds:
            tmin, tmax = thresholds[param].get("min"), thresholds[param].get("max")
            for cell in row:
                if cell.value is None:
                    continue
                try:
                    val = float(str(cell.value).replace(",", "."))
                except Exception:
                    continue
                if tmin is not None and tmax is not None:
                    if tmin <= val <= tmax:
                        cell.fill = green
                    else:
                        cell.fill = red

    wb.save(OUTPUT_FILE)
    print("[INFO] Colorisation Excel terminée")

# ---------------------------
# Graphiques
# ---------------------------
def plot_graphs():
    if not os.path.exists(OUTPUT_FILE):
        return
    df = pd.read_excel(OUTPUT_FILE)
    df = df.set_index("Paramètre")

    plt.ioff()  # mode non interactif

    for param in df.index:
        # Ne garder que les colonnes dont l'en-tête ressemble à une date
        date_cols = [
            c for c in df.columns
            if not pd.isna(pd.to_datetime(c, dayfirst=True, errors="coerce"))
        ]
        if not date_cols:
            continue
        series = df.loc[param, date_cols]
        # convertir en numérique, ignorer les cellules non numériques
        series = pd.to_numeric(series, errors="coerce").dropna()
        if len(series) > 1:
            # trier par date croissante
            idx_dates = pd.to_datetime(series.index, dayfirst=True, errors="coerce")
            series.index = idx_dates
            series = series.sort_index()

            plt.figure()
            plt.plot(series.index, series.values, marker='o')
            plt.title(f"Évolution de {param}")
            plt.xlabel("Date")
            plt.ylabel("Valeur")
            plt.grid(True)
            plt.tight_layout()
            safe_name = re.sub(r"[^A-Za-z0-9_\-]", "_", param)
            plt.savefig(f"graph_{safe_name}.png")
            plt.close()
    print("[INFO] Graphiques générés dans le dossier courant")

# ---------------------------
# Main
# ---------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true", help="Forcer ré-extraction")
    args = parser.parse_args()

    pdf_files = [f for f in os.listdir(PDF_FOLDER) if f.endswith(".pdf")]
    for pdf_file in pdf_files:
        pdf_path = os.path.join(PDF_FOLDER, pdf_file)
        date, results = extract_data_from_pdf(pdf_path)
        update_excel(pdf_path, date, results, force=args.force)

    colorize_outliers()
    plot_graphs()
