import os
import re
import json
import pdfplumber
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt
import argparse

load_dotenv()

# Valeurs par défaut si .env n'est pas défini
PDF_FOLDER = os.getenv("PDF_FOLDER", "pdfs")
OUTPUT_FILE = os.getenv("OUTPUT_FILE", "resultats.xlsx")
THRESHOLDS_FILE = os.getenv("THRESHOLDS_FILE", "seuils.json")
BANLIST_FILE = os.getenv("BANLIST_FILE", "banlist.txt")

# ---------------------------
# Liste blanche des paramètres biologiques
# ---------------------------
PARAMS_WHITELIST = [
    "Hématies", "Hémoglobine", "Hématocrite", "V.G.M.", "T.C.M.H.",
    "C.C.M.H.", "Leucocytes", "Polynucléaires neutrophiles",
    "Polynucléaires éosinophiles", "Polynucléaires basophiles",
    "Lymphocytes", "Monocytes", "Plaquettes",
    "Créatinine", "Phosphatase alcaline", "ASAT (Transaminases TGO)",
    "ALAT (Transaminases TGP)", "GGT (Gamma Glutamyl Transpeptidase)",
    "FIB-4"
]

# ---------------------------
# Extraction des données du PDF
# ---------------------------
def extract_data_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    # Date de prélèvement
    date_match = re.search(r"Prélevé le (\d{2}[-−]\d{2}[-−]\d{4})", text)
    if date_match:
        date = date_match.group(1).replace("−","-")
        text = text.split(date_match.group(0), 1)[1]  # texte après la date
    else:
        date = "inconnue"

    results = {}
    pattern = re.compile(
        r"^\s*([A-Za-zÀ-ÿ\s\-\(\)/]+?)\s+([\d,\.]+)\s*([a-zA-Zµ%/]+)?\s*(?:\(([^)]+)\))?"
    )

    # Charger banlist si présente
    banlist = []
    if BANLIST_FILE and os.path.exists(BANLIST_FILE):
        try:
            with open(BANLIST_FILE, "r", encoding="utf-8") as f:
                for raw in f:
                    s = raw.strip()
                    if not s or s.startswith("#"):
                        continue
                    banlist.append(s)
        except Exception:
            pass

    for line in text.split("\n"):
        line = line.strip()
        if not line or len(line) > 120:
            continue

        # 1) Essayer de parser les paramètres whitelistés (prioritaire sur banlist)
        parsed = False
        for param in PARAMS_WHITELIST:
            if line.startswith(param):
                match = pattern.match(line)
                if match:
                    val_str = match.group(2).replace(",", ".")
                    try:
                        value = float(val_str)
                        interval = match.group(4) or ""
                        low, high = None, None
                        if interval and "-" in interval:
                            parts = interval.replace("−","-").split("-")
                            try:
                                low = float(parts[0].replace(",", "."))
                                high = float(parts[1].replace(",", "."))
                            except:
                                pass
                        results[param] = {
                            "valeur": value,
                            "unité": match.group(3) or "",
                            "intervalle": interval,
                            "min": low,
                            "max": high
                        }
                        parsed = True
                    except ValueError:
                        pass
                break  # ne pas matcher plusieurs fois
        if parsed:
            continue

        # 2) Si non-parsable et la ligne contient un élément banni (insensible à la casse), ignorer
        line_low = line.lower()
        if any(b.lower() in line_low for b in banlist):
            continue

    return date, results

# ---------------------------
# Seuils: chargement / sauvegarde
# ---------------------------
def load_thresholds():
    thresholds = {}
    if not THRESHOLDS_FILE or not os.path.exists(THRESHOLDS_FILE):
        return thresholds

    _, ext = os.path.splitext(THRESHOLDS_FILE)
    try:
        if ext.lower() in [".json"]:
            with open(THRESHOLDS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            # normaliser {param: {min: x, max: y}}
            for k, v in data.items():
                try:
                    thresholds[k] = {
                        "min": None if v.get("min") is None else float(str(v.get("min")).replace(",", ".")),
                        "max": None if v.get("max") is None else float(str(v.get("max")).replace(",", ".")),
                    }
                except Exception:
                    continue
        elif ext.lower() in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
            df_thresh = pd.read_excel(THRESHOLDS_FILE, engine="openpyxl")
            for _, row in df_thresh.iterrows():
                try:
                    pname = str(row["Paramètre"]) if "Paramètre" in row else None
                    if not pname:
                        continue
                    vmin = row["min"] if "min" in row else None
                    vmax = row["max"] if "max" in row else None
                    thresholds[pname] = {
                        "min": None if vmin is None or pd.isna(vmin) else float(vmin),
                        "max": None if vmax is None or pd.isna(vmax) else float(vmax),
                    }
                except Exception:
                    continue
        else:
            print(f"[WARN] Extension de THRESHOLDS_FILE non supportée: {ext}")
    except Exception as e:
        print(f"[WARN] Impossible de lire THRESHOLDS_FILE : {e}")

    return thresholds


def save_thresholds(thresholds: dict):
    if not THRESHOLDS_FILE:
        return
    os.makedirs(os.path.dirname(THRESHOLDS_FILE) or ".", exist_ok=True)
    _, ext = os.path.splitext(THRESHOLDS_FILE)
    try:
        if ext.lower() == ".json":
            with open(THRESHOLDS_FILE, "w", encoding="utf-8") as f:
                json.dump(thresholds, f, ensure_ascii=False, indent=2)
        else:
            # Sauvegarde JSON par défaut même si extension non-json
            with open(THRESHOLDS_FILE + ".json", "w", encoding="utf-8") as f:
                json.dump(thresholds, f, ensure_ascii=False, indent=2)
            print(f"[INFO] Seuils enregistrés dans {THRESHOLDS_FILE + '.json'}")
    except Exception as e:
        print(f"[WARN] Échec d'enregistrement des seuils: {e}")

# ---------------------------
# Mise à jour Excel
# ---------------------------
def update_excel(pdf_path, date, results, force=False):
    if os.path.exists(OUTPUT_FILE):
        df = pd.read_excel(OUTPUT_FILE)
        if date in df.columns and not force:
            print(f"[INFO] Date {date} déjà extraite, passez --force pour ré-extraire")
            return
    else:
        df = pd.DataFrame()

    if "Paramètre" not in df.columns:
        df["Paramètre"] = list(results.keys())
    # Colonnes de base
    for base_col in ["Unité", "Min", "Max"]:
        if base_col not in df.columns:
            df[base_col] = None

    if date not in df.columns:
        df[date] = None

    # Charger seuils existants et les compléter à partir des PDF (si min/max présents)
    thresholds = load_thresholds()

    for param, info in results.items():
        if param not in df["Paramètre"].values:
            df.loc[len(df)] = [param] + [None]*(len(df.columns)-1)
        df.loc[df["Paramètre"]==param, date] = info["valeur"]

        # Enregistrer automatiquement un seuil si manquant et détecté
        if info.get("min") is not None and info.get("max") is not None:
            if param not in thresholds:
                thresholds[param] = {"min": info["min"], "max": info["max"]}

        # Renseigner Unité/Min/Max dans l'Excel si disponibles et absents
        if info.get("unité") is not None:
            df.loc[df["Paramètre"] == param, "Unité"] = info["unité"]
        if info.get("min") is not None:
            df.loc[df["Paramètre"] == param, "Min"] = info["min"]
        if info.get("max") is not None:
            df.loc[df["Paramètre"] == param, "Max"] = info["max"]

    # Sauvegarder les seuils mis à jour (si quelque chose a été ajouté)
    save_thresholds(thresholds)

    # Réordonner les colonnes de dates par ordre chronologique
    def _is_date_label(c):
        if c == "Paramètre":
            return False
        if c in ("Unité", "Min", "Max"):
            return False
        return not pd.isna(pd.to_datetime(c, dayfirst=True, errors="coerce"))

    date_cols = [c for c in df.columns if _is_date_label(c)]
    non_date_cols = [c for c in df.columns if c not in date_cols and c != "Paramètre"]
    # tri par date croissante
    sorted_dates = sorted(date_cols, key=lambda x: pd.to_datetime(x, dayfirst=True, errors="coerce"))
    ordered_cols = ["Paramètre", "Unité", "Min", "Max"] + sorted_dates + non_date_cols
    df = df.reindex(columns=ordered_cols)

    df.to_excel(OUTPUT_FILE, index=False)
    print(f"[INFO] {os.path.basename(pdf_path)} -> {len(results)} résultats extraits pour la date {date}")

# ---------------------------
# Colorisation Excel
# ---------------------------
def colorize_outliers():
    if not os.path.exists(OUTPUT_FILE):
        return

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    green = PatternFill(start_color="AAFFAA", end_color="AAFFAA", fill_type="solid")
    red = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")

    # Charger les seuils (JSON ou Excel)
    thresholds = load_thresholds()

    # Repérer colonnes
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
    idx_param = headers.index("Paramètre") + 1 if "Paramètre" in headers else 1
    idx_min = headers.index("Min") + 1 if "Min" in headers else None
    idx_max = headers.index("Max") + 1 if "Max" in headers else None
    # Colonnes de dates
    date_cols_idx = []
    for i, h in enumerate(headers, start=1):
        if h in ("Paramètre", "Unité", "Min", "Max"):
            continue
        try:
            if not pd.isna(pd.to_datetime(h, dayfirst=True, errors="coerce")):
                date_cols_idx.append(i)
        except Exception:
            pass

    for r in range(2, ws.max_row+1):
        param = ws.cell(row=r, column=idx_param).value
        # seuils préférentiels: colonnes Min/Max de la feuille, sinon JSON
        tmin = tmax = None
        if idx_min and idx_max:
            vmin = ws.cell(row=r, column=idx_min).value
            vmax = ws.cell(row=r, column=idx_max).value
            try:
                tmin = None if vmin in (None, "") else float(str(vmin).replace(",", "."))
                tmax = None if vmax in (None, "") else float(str(vmax).replace(",", "."))
            except Exception:
                tmin = tmax = None
        if (tmin is None or tmax is None) and param in thresholds:
            tmin = thresholds[param].get("min")
            tmax = thresholds[param].get("max")

        if tmin is None or tmax is None:
            continue
        for c in date_cols_idx:
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            try:
                val = float(str(cell.value).replace(",", "."))
            except Exception:
                continue
            cell.fill = green if (tmin <= val <= tmax) else red

    for row in ws.iter_rows(min_row=2, min_col=2):
        param = ws.cell(row=row[0].row, column=1).value
        if param in thresholds:
            tmin, tmax = thresholds[param].get("min"), thresholds[param].get("max")
            for cell in row:
                if cell.value is None:
                    continue
                try:
                    val = float(str(cell.value).replace(",", "."))
                except Exception:
                    continue
                if tmin is not None and tmax is not None:
                    if tmin <= val <= tmax:
                        cell.fill = green
                    else:
                        cell.fill = red

    wb.save(OUTPUT_FILE)
    print("[INFO] Colorisation Excel terminée")

# ---------------------------
# Graphiques
# ---------------------------
def insert_charts_into_excel():
    if not os.path.exists(OUTPUT_FILE):
        return
    # Charger workbook et feuille de données
    wb = load_workbook(OUTPUT_FILE)
    ws_data = wb.active

    # Trouver/Créer la feuille "Graphiques"
    sheet_name = "Graphiques"
    if sheet_name in wb.sheetnames:
        # Supprimer et recréer pour éviter doublons
        del wb[sheet_name]
    ws_charts = wb.create_sheet(title=sheet_name)

    # Repérer les colonnes de dates (en-têtes ligne 1)
    headers = [ws_data.cell(row=1, column=col).value for col in range(1, ws_data.max_column+1)]
    try:
        date_cols_idx = [
            i+1 for i, h in enumerate(headers)
            if h not in ("Paramètre", "Unité", "Min", "Max")
            and not pd.isna(pd.to_datetime(h, dayfirst=True, errors="coerce"))
        ]
    except Exception:
        date_cols_idx = [i+1 for i, h in enumerate(headers) if h not in ("Paramètre", "Unité", "Min", "Max")]

    # Pour chaque paramètre (ligne >=2), créer un graphique si au moins 2 valeurs
    chart_row = 1
    chart_col = 1
    charts_per_row = 2
    for r in range(2, ws_data.max_row+1):
        param_name = ws_data.cell(row=r, column=1).value
        if not param_name:
            continue

        # Compter valeurs numériques présentes
        numeric_count = 0
        for c in date_cols_idx:
            v = ws_data.cell(row=r, column=c).value
            try:
                if v is not None and str(v) != "" and float(str(v).replace(",", ".")) == float(str(v).replace(",", ".")):
                    numeric_count += 1
            except Exception:
                pass
        if numeric_count < 2:
            continue

        # Construire chart
        chart = LineChart()
        chart.title = f"Évolution de {param_name}"
        chart.y_axis.title = "Valeur"
        chart.x_axis.title = "Date"

        # Données
        data_ref = Reference(ws_data, min_col=min(date_cols_idx), max_col=max(date_cols_idx), min_row=r, max_row=r)
        cat_ref = Reference(ws_data, min_col=min(date_cols_idx), max_col=max(date_cols_idx), min_row=1, max_row=1)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cat_ref)

        # Position dans la feuille charts
        cell_addr = ws_charts.cell(row=chart_row, column=chart_col).coordinate
        ws_charts.add_chart(chart, cell_addr)

        # Avancer position (grille 2 colonnes)
        if chart_col == 1:
            chart_col = 9  # environ 8 colonnes d'écart
        else:
            chart_col = 1
            chart_row += 15  # descendre d'un bloc

    wb.save(OUTPUT_FILE)
    print("[INFO] Graphiques intégrés dans le fichier Excel (onglet 'Graphiques')")

# ---------------------------
# Main
# ---------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true", help="Forcer ré-extraction")
    args = parser.parse_args()

    pdf_files = [f for f in os.listdir(PDF_FOLDER) if f.endswith(".pdf")]
    for pdf_file in pdf_files:
        pdf_path = os.path.join(PDF_FOLDER, pdf_file)
        date, results = extract_data_from_pdf(pdf_path)
        update_excel(pdf_path, date, results, force=args.force)

    colorize_outliers()
    insert_charts_into_excel()
